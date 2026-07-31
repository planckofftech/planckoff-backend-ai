"""The master sheet is where every source will eventually meet.

Its promise is that a blank cell means "nobody has told us yet". A door
schedule can answer roughly a third of the columns; filling the rest with
plausible defaults would destroy the only signal the sheet has.
"""

import io

import openpyxl
import pytest

from app.core.master_sheet import COLUMNS, build_workbook, master_row
from app.pipeline import extract
from app.schemas import DoorRow, ExtractionResult, ExtractionMethod


def _result(rows: list[DoorRow]) -> ExtractionResult:
    return ExtractionResult(
        method=ExtractionMethod.DETERMINISTIC_RULED, pages_scanned=1,
        row_count=len(rows), rows=rows,
    )


def test_template_column_set_is_the_sheets_own():
    assert len(COLUMNS) == 41
    assert COLUMNS[0] == "DOOR TAG"
    assert COLUMNS[10] == "WIDTH"
    assert COLUMNS[24] == "FRAME MATERIAL"
    assert COLUMNS[-2] == "HARDWARE SET"


def test_columns_no_source_answered_stay_empty():
    """Not 'N/A', not a guess. A reader must be able to tell an unknown from a
    measurement."""
    values = master_row(DoorRow(door_tag="101A", door_width="3' - 0\""))

    for column in ("BUILDING TAG", "DOOR CORE", "FRAME ANCHOR", "PREHUNG",
                   "WALL TYPE", "HAND OF OPENINGS"):
        assert values.get(column, "") == "", f"{column} was invented"


def test_direction_and_single_location_both_become_door_location():
    both = master_row(DoorRow(from_space="HALL", to_space="OFFICE"))
    assert both["DOOR LOCATION"] == "HALL to OFFICE"

    single = master_row(DoorRow(extra={"location": "STAIR A (CELLAR)"}))
    assert single["DOOR LOCATION"] == "STAIR A (CELLAR)"


def test_interior_exterior_is_only_stated_when_the_drawing_says_so():
    assert master_row(DoorRow(from_space="EXTERIOR",
                              to_space="LOBBY"))["INTERIOR/EXTERIOR"] == "EXTERIOR"
    # Silence is not "INTERIOR" -- that would be a guess presented as a reading.
    assert master_row(DoorRow(from_space="HALL",
                              to_space="OFFICE")).get("INTERIOR/EXTERIOR", "") == ""


def test_leaf_count_read_from_a_pair_width():
    assert master_row(DoorRow(door_width="2*3' - 0\""))["LEAF COUNT"] == "2"
    assert master_row(DoorRow(door_width="3' - 0\"")).get("LEAF COUNT", "") == ""


def test_thickness_is_found_under_whatever_the_firm_calls_it():
    for key in ("thk", "door_thickness", "panel_thk"):
        values = master_row(DoorRow(extra={key: '1 3/4"'}))
        assert values["THICKNESS"] == '1 3/4"', key


def test_workbook_keeps_the_two_header_rows_and_bands():
    xlsx, stats = build_workbook(_result([DoorRow(door_tag="1", door_width="3'")]))
    ws = openpyxl.load_workbook(io.BytesIO(xlsx)).active

    assert ws.cell(1, 1).value == "BASIC INFORMATION"
    assert ws.cell(1, 15).value == "DOOR"
    assert ws.cell(1, 25).value == "FRAME"
    assert ws.cell(1, 40).value == "HARDWARE"
    assert [ws.cell(2, c).value for c in range(1, 42)] == COLUMNS
    # Data starts on row 3, under the two header rows.
    assert ws.cell(3, 1).value == "1"
    assert stats.rows == 1


@pytest.mark.asyncio
async def test_every_schedule_on_the_sheet_reaches_the_master(multi_schedule_bytes):
    """Three schedules on one drawing must all land in the master sheet, not
    just the largest."""
    result = await extract(multi_schedule_bytes, allow_ai=False)
    expected = sum(t.row_count for t in result.tables)

    xlsx, stats = build_workbook(result)
    ws = openpyxl.load_workbook(io.BytesIO(xlsx)).active

    assert stats.rows == expected > result.row_count
    assert ws.max_row == 2 + expected


@pytest.mark.asyncio
async def test_endpoint_maps_rows_and_never_re_extracts(ellis_p21_bytes, monkeypatch):
    """The master sheet takes an extraction, not a PDF.

    Taking a PDF meant the document was read twice for one sheet: twice the
    wait, twice the AI cost, and a second chance to fail after the first read
    had already succeeded.
    """
    import json as _json

    from fastapi.testclient import TestClient

    import app.pipeline as pipeline
    from app.config import get_settings
    from app.main import app

    result = await extract(ellis_p21_bytes, allow_ai=False)
    payload = _json.loads(result.model_dump_json())

    async def explode(*_a, **_kw):
        raise AssertionError("master sheet re-extracted the document")

    monkeypatch.setattr(pipeline, "extract", explode)

    with TestClient(app) as client:
        response = client.post(
            "/api/v1/master-sheet?preview=true",
            headers={"X-API-Key": get_settings().api_key}, json=payload,
        )

    assert response.status_code == 200
    assert response.json()["row_count"] == 23


@pytest.mark.asyncio
async def test_preview_and_download_are_built_from_the_same_rows(ellis_p21_bytes):
    """What the screen shows must be what the spreadsheet contains -- a preview
    that disagrees with the download is worse than no preview."""
    import json as _json

    from fastapi.testclient import TestClient

    from app.config import get_settings
    from app.main import app

    result = await extract(ellis_p21_bytes, allow_ai=False)
    payload = _json.loads(result.model_dump_json())
    headers = {"X-API-Key": get_settings().api_key}

    with TestClient(app) as client:
        shown = client.post("/api/v1/master-sheet?preview=true",
                            headers=headers, json=payload).json()
        downloaded = client.post("/api/v1/master-sheet",
                                 headers=headers, json=payload)

    assert shown["columns"] == COLUMNS
    ws = openpyxl.load_workbook(io.BytesIO(downloaded.content)).active
    assert ws.max_row - 2 == shown["row_count"] == len(shown["rows"])

    for offset, row in enumerate(shown["rows"]):
        for index, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(3 + offset, index).value or ""
            assert cell == row[name], f"row {offset} column {name} disagrees"


@pytest.mark.asyncio
async def test_real_extraction_fills_the_columns_it_can(ellis_p21_bytes):
    result = await extract(ellis_p21_bytes, allow_ai=False)
    xlsx, stats = build_workbook(result)
    ws = openpyxl.load_workbook(io.BytesIO(xlsx)).active

    assert stats.rows == 23
    header = {name: i + 1 for i, name in enumerate(COLUMNS)}
    row = {n: ws.cell(4, i).value for n, i in header.items()}  # door 1

    assert row["DOOR TAG"] == "1"
    assert row["DOOR LOCATION"] == "RECEPTION to EXTERIOR"
    assert row["WIDTH"] == "6' - 0\""
    assert row["DOOR MATERIAL"] == "STOREFRONT"
    assert row["HARDWARE SET"] == "1"
    assert row["INTERIOR/EXTERIOR"] == "EXTERIOR"
    assert row["QUANTITY"] == "1"
    assert "DOOR CORE" in stats.empty_columns

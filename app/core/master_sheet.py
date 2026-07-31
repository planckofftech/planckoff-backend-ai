"""Fill the master door format sheet from an extraction.

The master sheet is the one place per project where every source meets: door
schedules, hardware documents, pricing. This module handles the first source --
a door schedule -- and deliberately leaves every other column blank rather than
guessing, so a reader can tell "nobody has told us yet" from "measured".
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas import DoorRow, ExtractionResult

# Column bands as printed on the template, in order.
BANDS: list[tuple[str, int]] = [
    ("BASIC INFORMATION", 1), ("DOOR", 15), ("FRAME", 25), ("HARDWARE", 40),
]

COLUMNS: list[str] = [
    "DOOR TAG", "BUILDING TAG", "BUILDING LOCATION", "DOOR LOCATION", "QUANTITY",
    "HAND OF OPENINGS", "DOOR OPERATION", "LEAF COUNT", "INTERIOR/EXTERIOR",
    "EXCLUDE REASON", "WIDTH", "HEIGHT", "THICKNESS", "FIRE RATING",
    "DOOR MATERIAL", "DOOR ELEVATION TYPE", "DOOR CORE", "DOOR FACE", "DOOR EDGE",
    "DOOR GUAGE", "DOOR FINISH", "STC RATING", "DOOR UNDERCUT",
    "DOOR INCLUDE/EXCLUDE", "FRAME MATERIAL", "WALL TYPE", "THROAT THICKNESS",
    "FRAME ANCHOR", "BASE ANCHOR", "NO OF ANCHOR", "FRAME PROFILE",
    "FRAME ELEVATION TYPE", "FRAME ASSEMBLY", "FRAME GUAGE", "FRAME FINISH",
    "PREHUNG", "FRAME HEAD", "CASING", "FRAME INCLUDE/EXCLUDE", "HARDWARE SET",
    "HARDWARE INCLUDE/EXCLUDE",
]

# Canonical field -> master column.
_DIRECT: dict[str, str] = {
    "door_tag": "DOOR TAG",
    "door_width": "WIDTH",
    "door_height": "HEIGHT",
    "fire_rating": "FIRE RATING",
    "door_material": "DOOR MATERIAL",
    "door_type": "DOOR ELEVATION TYPE",
    "door_finish": "DOOR FINISH",
    "frame_material": "FRAME MATERIAL",
    "frame_finish": "FRAME FINISH",
    "hw_set": "HARDWARE SET",
}

# Columns a door schedule sometimes carries under its own name. First key that
# the row actually has wins, so one list serves every firm's naming.
_FROM_EXTRA: dict[str, tuple[str, ...]] = {
    "THICKNESS": ("thk", "door_thickness", "thickness", "panel_thk",
                  "door_as_applicable_thk"),
    "STC RATING": ("stc_rating",),
    "FRAME ELEVATION TYPE": ("frame_type",),
    "FRAME GUAGE": ("frame_gauge", "gauge"),
    "FRAME HEAD": ("detail_reference_head", "head", "frame_head_detail"),
    "WALL TYPE": ("wall_type",),
    "DOOR GUAGE": ("door_gauge",),
}

_EXTERIOR = re.compile(r"\bEXT(ERIOR)?\b", re.I)
# "2*3' - 0"" or "2 x 3'-0"" means a pair.
_LEAF_COUNT = re.compile(r"^\s*(\d+)\s*[*x]\s*\d", re.I)

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_BAND_FILL = PatternFill("solid", fgColor="2F5597")
_THIN = Side(style="thin", color="B4C6E7")


@dataclass(slots=True)
class MasterSheetStats:
    rows: int
    filled_columns: list[str]
    empty_columns: list[str]


def _location(row: DoorRow) -> str:
    """Where the door is. A schedule that names both sides gives a direction;
    one that names a single location gives that."""
    parts = [p for p in (row.from_space, row.to_space) if p]
    if parts:
        return " to ".join(parts)
    for key in ("location", "lock_function", "room"):
        if row.extra.get(key):
            return row.extra[key]
    return ""


def _interior_exterior(row: DoorRow) -> str:
    """Only stated when the drawing says so. Silence is not 'INTERIOR'."""
    for value in (row.from_space, row.to_space, _location(row)):
        if value and _EXTERIOR.search(value):
            return "EXTERIOR"
    return ""


def _leaf_count(row: DoorRow) -> str:
    match = _LEAF_COUNT.match(row.door_width or "")
    return match.group(1) if match else ""


def _first_extra(row: DoorRow, keys: tuple[str, ...]) -> str:
    for key in keys:
        value = row.extra.get(key)
        if value:
            return value
    return ""


def master_row(row: DoorRow) -> dict[str, str]:
    """One door as master-sheet columns. Unknown columns stay absent, never
    filled with a plausible-looking default."""
    out: dict[str, str] = {}
    data = row.model_dump()

    for field, column in _DIRECT.items():
        if data.get(field):
            out[column] = data[field]

    for column, keys in _FROM_EXTRA.items():
        value = _first_extra(row, keys)
        if value:
            out[column] = value

    if location := _location(row):
        out["DOOR LOCATION"] = location
    if side := _interior_exterior(row):
        out["INTERIOR/EXTERIOR"] = side
    if leaves := _leaf_count(row):
        out["LEAF COUNT"] = leaves

    # A schedule lists each opening once unless it says otherwise.
    out.setdefault("QUANTITY", "1")
    return out


def build_workbook(result: ExtractionResult, *, source_name: str = ""
                   ) -> tuple[bytes, MasterSheetStats]:
    """The master sheet as .xlsx, with every extracted door as a row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    band_at = {index: name for name, index in BANDS}
    for index, name in enumerate(COLUMNS, start=1):
        band = ws.cell(1, index)
        if index in band_at:
            band.value = band_at[index]
            band.font = Font(bold=True, color="FFFFFF")
            band.fill = _BAND_FILL
        head = ws.cell(2, index)
        head.value = name
        head.font = Font(bold=True, color="FFFFFF", size=9)
        head.fill = _HEADER_FILL
        head.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        head.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
        ws.column_dimensions[head.column_letter].width = max(12, min(26, len(name) + 4))

    ws.freeze_panes = "A3"

    # Every schedule on the sheet, in the order they appear on the drawing.
    tables = result.tables or []
    rows = [r for t in tables for r in t.rows] if tables else list(result.rows)

    filled: set[str] = set()
    for offset, row in enumerate(rows):
        values = master_row(row)
        filled.update(k for k, v in values.items() if v)
        for index, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(3 + offset, index)
            cell.value = values.get(name, "")
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    ws.auto_filter.ref = f"A2:{ws.cell(2, len(COLUMNS)).column_letter}{2 + len(rows)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()

    stats = MasterSheetStats(
        rows=len(rows),
        filled_columns=[c for c in COLUMNS if c in filled],
        empty_columns=[c for c in COLUMNS if c not in filled],
    )
    _ = source_name
    return buffer.getvalue(), stats
